import openpyxl

class ContactBook:
    # def __init__(self, file_name="contacts.xlsx"):      #OnlineUseComment
    def __init__(self):                                 #LocalUseComment
        # self.file_name = file_name                      #OnlineUseComment
        self.contacts = {}
        # self.load_contacts()                            #OnlineUseComment

    def add_contact(self, name, phone, email, address):
        self.contacts[name] = {"phone": phone, "email": email, "address": address}
        # self.save_contacts()                            #OnlineUseComment
        
    def view_contacts(self):
        if not self.contacts:
            print("Contact list is empty.")
            return
    
        print("\nContact List:-\n")
        for name, details in self.contacts.items():
            print(f"Name: {name}")
            print(f"Phone: {details['phone']}")
            print("-" * 20)
    
    def search_contacts(self, search_term):
        found_contacts = []
        for name, details in self.contacts.items():
            if search_term.lower() in name.lower() or search_term in details['phone']:
                found_contacts.append((name, details))

        if not found_contacts:
            print("No matching contacts found.")
            return
        
        print("\n")
        for name, details in found_contacts:
            print(f"Name: {name}")
            print(f"Phone: {details['phone']}")
            print(f"Email: {details['email']}")
            print(f"Address: {details['address']}")
            print("-" * 20)


    def update_contact(self, name, phone, email, address):
        if name in self.contacts:
            self.contacts[name] = {"phone": phone, "email": email, "address": address}
            print(f"Contact {name} updated successfully.")
        else:
            print(f"Contact {name} not found.")

    def delete_contact(self, name):
        if name in self.contacts:
            del self.contacts[name]
            print(f"Contact {name} deleted successfully.")
        else:
            print(f"Contact {name} not found.")
    
    def save_contacts(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Phone", "Email", "Address"])
        
        for name, details in self.contacts.items():
            sheet.append([name, details["phone"], details["email"], details["address"]])
        
        workbook.save(self.file_name)

    def load_contacts(self):
        try:
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, phone, email, address = row
                self.contacts[name] = {"phone": phone, "email": email, "address": address}
        except FileNotFoundError:
            pass


#Main()_Body
contact_book = ContactBook()

while True:
    print("\n\tMAIN MENU")
    print("1. Add Contact")
    print("2. View Contact List")
    print("3. Search Contact")
    print("4. Update Contact")
    print("5. Delete Contact")
    print("6. Exit")

    choice = input("\nEnter your choice: ")

    if choice == "1":
        name = input("Enter Name: ")
        phone = input("Enter Phone: ")
        email = input("Enter Email: ")
        address = input("Enter Address: ")
        contact_book.add_contact(name, phone, email, address)

    elif choice == "2":
        contact_book.view_contacts()

    elif choice == "3":
        search_term = input("Enter search term: ")
        contact_book.search_contacts(search_term)

    elif choice == "4":
        name = input("Enter Name of the contact to update: ")
        phone = input("Enter new Phone: ")
        email = input("Enter new Email: ")
        address = input("Enter new Address: ")
        contact_book.update_contact(name, phone, email, address)

    elif choice == "5":
        name = input("Enter Name of the contact to delete: ")
        contact_book.delete_contact(name)

    elif choice == "6":
        print("Exiting the Contact Book application.")
        break
        
    else:
        print("Invalid choice. Please select a valid option.")

